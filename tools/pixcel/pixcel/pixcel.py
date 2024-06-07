import os
import sys
import openpyxl
import argparse
from PIL import Image
from pathlib import Path
from openpyxl.styles import PatternFill

def ex2pix(excel_path, num_rows, num_columns, output_path=None):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    image = Image.new('RGB', (num_columns, num_rows))
    pixels = image.load()
    
    if output_folder is None:
        output_folder = 'output'
        os.makedirs(output_folder, exist_ok=True)

    for y in range(1, num_rows + 1):
        for x in range(1, num_columns + 1):
            cell = sheet.cell(row=y, column=x)
            fill = cell.fill

            if fill.fgColor.type == 'rgb':
                color = fill.fgColor.rgb
                r = int(color[2:4], 16)
                g = int(color[4:6], 16)
                b = int(color[6:8], 16)
                pixels[x - 1, y - 1] = (r, g, b)
            else:
                pixels[x - 1, y - 1] = (255, 255, 255)

    image.save(output_path)


def pix2ex(path_to_img, num_rows, num_cols, output_folder=None):
    image = Image.open(image_path)
    image = resize_image(image, (num_rows, num_cols))
    pixels = image.load()
    width, height = image.size

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if output_folder is None:
        output_folder = 'output'
        os.makedirs(output_folder, exist_ok=True)

    for y in range(height):
        for x in range(width):
            r, g, b = pixels[x, y]
            fill = PatternFill(start_color=f'{r:02X}{g:02X}{b:02X}',
                               end_color=f'{r:02X}{g:02X}{b:02X}',
                               fill_type='solid')
            cell = sheet.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    output_path = os.path.join(output_folder, os.path.splitext(os.path.basename(image_path))[0] + '.xlsx')
    workbook.save(output_path)

    return output_path
    
def main():
    parser = argparse.ArgumentParser(description="Convert between Excel sheets and images.")

    subparsers = parser.add_subparsers(dest="command", help="Choose a command")

    ex2pix_parser = subparsers.add_parser("ex2pix", help="Convert Excel sheet to image")
    ex2pix_parser.add_argument("excel_path", type=str, help="Path to the Excel file")
    ex2pix_parser.add_argument("num_rows", type=int, help="Number of rows in the image")
    ex2pix_parser.add_argument("num_columns", type=int, help="Number of columns in the image")
    ex2pix_parser.add_argument("--output_path", "-o", type=str, help="Output path for the image")

    pix2ex_parser = subparsers.add_parser("pix2ex", help="Convert image to Excel sheet")
    pix2ex_parser.add_argument("image_path", type=str, help="Path to the image")
    pix2ex_parser.add_argument("num_rows", type=int, help="Number of rows in the Excel sheet")
    pix2ex_parser.add_argument("num_columns", type=int, help="Number of columns in the Excel sheet")
    pix2ex_parser.add_argument("--output_folder", "-o", type=str, help="Output folder for the Excel sheet")

    args = parser.parse_args()

    if args.command == "ex2pix":
        output_path = args.output_path
        if not output_path:
            output_path = Path(args.excel_path).with_suffix(".png")
        ex2pix(args.excel_path, args.num_rows, args.num_columns, output_path)
        print(f"Pixel art image saved to: {output_path}")

    elif args.command == "pix2ex":
        output_folder = args.output_folder
        if not output_folder:
            output_folder = "output"
        pix2ex(args.image_path, args.num_rows, args.num_columns, output_folder)
        print(f"Excel sheet saved to: {output_folder}")

if __name__ == "__main__":
    main()
